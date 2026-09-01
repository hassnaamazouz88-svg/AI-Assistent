from openpyxl import load_workbook


def extract_xlsx(file_path: str) -> dict:
    """
    Extrait le contenu structuré d'un fichier XLSX, sans le
    transformer en texte aplati — ce flattening est délégué à
    text_builder.py, juste avant le chunking.

    Retourne :
    {
        "sheets": {
            "NomFeuille": [
                ["cellule1", "cellule2", ...],  # ligne 1
                ["cellule1", "cellule2", ...],  # ligne 2
            ]
        },
        "nb_sheets": 1
    }
    """

    workbook = load_workbook(
        filename=file_path,
        data_only=True
    )

    sheets = {}

    for sheet_name in workbook.sheetnames:

        worksheet = workbook[sheet_name]

        rows = []

        for row in worksheet.iter_rows(values_only=True):

            values = [
                str(value).strip() if value is not None else ""
                for value in row
            ]

            if any(values):
                rows.append(values)

        sheets[sheet_name] = rows

    return {
        "sheets": sheets,
        "nb_sheets": len(workbook.sheetnames)
    }